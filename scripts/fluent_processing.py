
from __future__ import annotations
import os
import numpy as np
import subprocess
import sys
import time
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import shutil
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from scripts.ui import FluentProcessingUI


rows: int = [11,    # Drag Frontwing
             12,    # Drag Sidepod
             13,    # Drag Rearwing
             15,    # Drag Net
             11,    # Df Frontwing
             12,    # Df Sidepod
             13,    # Df Rearwing
             15,    # Df Net
             18     # Moment around front axis
             ]

cols: int = [3,     # Drag Frontwing
             3,     # Drag Sidepod
             3,     # Drag Rearwing
             3,     # Drag Net
             8,     # Df Frontwing
             8,     # Df Sidepod
             8,     # Df Rearwing
             8,     # Df Net
             8      # Moment around front axis
             ]


class FluentPostProcesser():
    def __init__(self, fluent_exe_path: Path, case_folder_path: Path, ui: FluentProcessingUI):
        self.fluent_exe_path = fluent_exe_path
        self.ui = ui
        self.create_folder_struct(case_folder_path)
        self.progress_flag = False

    def create_folder_struct(self, case_folder_path):
        self.work_dir = Path(case_folder_path)
        self.case_file_path = list(case_folder_path.rglob("*.cas.h5"))[0]
        self.out_dir = self.work_dir / "processed"
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.jou_path = Path(os.path.abspath(r"data\v0.1_sequence.jou"))
        self.images_dir = self.out_dir / "images"
        self.images_dir.mkdir(parents=True, exist_ok=True)
        self.front_vel_dir = self.images_dir / "front_vel"
        self.front_vel_dir.mkdir(parents=True, exist_ok=True)
        self.front_pr_dir = self.images_dir / "front_pressure"
        self.front_pr_dir.mkdir(parents=True, exist_ok=True)
        self.side_vel_dir = self.images_dir / "side_vel"
        self.side_vel_dir.mkdir(parents=True, exist_ok=True)
        self.side_pr_dir = self.images_dir / "side_pressure"
        self.side_pr_dir.mkdir(parents=True, exist_ok=True)
        self.forces_dir = self.out_dir / "forces"
        self.forces_dir.mkdir(parents=True, exist_ok=True)

    def create_jou_content(self):
        # --- Parameters of sequence ---
        start_z = 0.0
        end_z = 0.760  # 740 mm
        num_images_y = 30
        num_images_x = 45
        z_positions = np.linspace(start_z, end_z, num_images_y)

        start_x = -1.250
        end_x = 1.850 
        x_positions = np.linspace(start_x, end_x, num_images_x)
        # --- creating jou content ---


        jou_content = f"""; 2047
            /file/read-case-data "{self.case_file_path.as_posix()}"
            /file/set-batch-options no yes yes no
            /display/set-window 1
            /views/camera/projection orthographic
            /views/apply-mirror-planes symmetry ()
            
            /report/forces/wall-forces yes 0 0 1 yes df.csv
            /report/forces/wall-forces yes 1 0 0 yes drag.csv

            /display/surface/iso-surface velocity velo_iso () () 5 ()
            /display/set/contours surfaces velo_iso ()
            /display/contour/pressure -600 100
            /views/auto-scale   
            /views/camera/target 0.5 0.1 0
            /views/camera/position 0.5 0.1 -1
            /views/camera/up-vector 0 1 0
            /views/camera/zoom-camera 1.5
            /display/save-picture "Processed/images/bottom_iso.png"

            /display/display/surface-mesh symmetry()
            /views/camera/target 0.4 0 1
            /views/camera/position 0.4 1 1
            /views/camera/up-vector 0 0 1   
            """ 

        for i, z in enumerate(z_positions):
            s_name = f"plane_z_{i:02d}"
            img_vel = f"Processed/Images/side_vel/side_vel_{i:02d}.png"
            img_pressure = f"Processed/Images/side_pressure/side_pressure_{i:02d}.png"
            
            jou_content += f"""
                ; --- Image {i+1}/{num_images_y} at Z = {z:.4f} ---
                /surface/plane-surface {s_name} z {z:.4f}
                /display/set/contours surfaces {s_name} ()
                /display/contour/velocity-magnitude 0 35   
                /display/save-picture "{img_vel}"
                /display/contour/total-pressure -300 350
                /display/save-picture "{img_pressure}"
                /surface/delete {s_name}
                """

        jou_content += f"""
            views/apply-mirror-planes symmetry ()
            /display/display/surface-mesh Inlet()
            /views/auto-scale
            /views/camera/target 0 0 0.8
            /views/camera/position 1 0 0.8
            /views/camera/up-vector 0 0 1
            /views/camera/zoom-camera 4
            """       

        for i, y in enumerate(x_positions):
            s_name = f"plane_x_{i:02d}"
            img_vel = f"Processed/Images/front_vel/front_vel_{i:02d}.png"
            img_pressure = f"Processed/Images/front_pressure/front_pressure_{i:02d}.png"
        
            
            jou_content += f"""
                ; --- Image {i+1}/{num_images_x} at x = {y:.4f} ---
                /surface/plane-surface {s_name} y {y:.4f}
                /display/set/contours surfaces {s_name} ()
                /display/contour/velocity-magnitude 0 35
                /display/save-picture "{img_vel}"
                /display/contour/total-pressure -300 300
                /display/save-picture "{img_pressure}"
                /surface/delete {s_name}
                """
            
        jou_content += "\n/exit yes"
        
        self.jou_path.write_text(jou_content, encoding="utf-8")

    def run_jou_file(self, timeout_s = 2000):
        print("Running Fluent (Mode Batch)...")

        with open(self.jou_path) as f:
            first_line = f.readline().strip()
            # remove a possible leading ';' or '#' used as comment
            first_line = first_line.lstrip(';#').strip()
            n_lines = int(first_line)

        cmd = [str(self.fluent_exe_path), "3d", "-t8", "-gu", "-i", self.jou_path]
        start = time.time()
        proc = subprocess.Popen(
            cmd,
            cwd=str(self.work_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            universal_newlines=True,
            )
        assert proc.stdout is not None
        try:
            for line_counter, line in enumerate(proc.stdout):
                if line_counter == 0:
                    self.jou_progress(n_lines, line_counter)
                    continue
                print(line, end="")
                sys.stdout.flush()
                self.jou_progress(n_lines, line_counter)
                if time.time() - start > timeout_s:
                    proc.kill()
                    raise TimeoutError(f"Process was longer than {timeout_s}s.")
        finally:
            try:
                proc.stdout.close()
            except Exception:
                pass
        rc = proc.wait()
        if rc == 0:
            print(f"\n Images saved in: {self.out_dir}")
        else:
            print(f"\n Error occoured in process: (Code {rc})")

    def jou_progress(self, n_lines, line_counter):
        if self.progress_flag == False:
            self.progress_old = 0
            self.progress = 0
            self.progress_flag = True
        else:
            self.progress_old = self.progress
        self.progress = round(line_counter/n_lines * 100)
        if self.progress != self.progress_old:
            self.ui.show_progress(self.progress)
            
    def get_excel_data(self):
        source = self.work_dir / "df.csv"
        destination = self.forces_dir
        try:
            shutil.copy2(source, destination)
            os.remove(source)
        except:
            print("Couldn't find df.csv")
        source = self.work_dir / "drag.csv"
        try:
            shutil.copy2(source, destination)
            os.remove(source)
        except:
            print("Couldn't find drag.csv")

        df = pd.read_csv(
            self.forces_dir / "df.csv",
            skiprows=19,
            sep=r"\s+",
            engine="python"
            )

        drag = pd.read_csv(
            self.forces_dir / "drag.csv",     # chemin vers ton fichier
            skiprows=19,        # ignore les 19 premières lignes
            sep=r"\s+",         # 1 ou plusieurs espaces comme séparateur
            engine="python"     # nécessaire pour les regex
            )  

        df_list = pd.to_numeric(df["Total"], errors="coerce").tolist()
        drag_list = pd.to_numeric(drag["Total"], errors="coerce").tolist()

        idx = {
        "chassis": 1,
        "front-wheel": 3,
        "front-wing": 4,
        "rear-wheel": 5,
        "rear-wing": 6,
        "sidepod": 8,
        }

        drag_fw = float(drag_list[idx["front-wing"]])
        drag_sp = float(drag_list[idx["sidepod"]])
        drag_rw = float(drag_list[idx["rear-wing"]])
        drag_rwh = float(drag_list[idx["rear-wheel"]])
        drag_fwh = float(drag_list[idx["front-wheel"]])
        drag_ch = float(drag_list[idx["chassis"]])
        drag_net = drag_fw + drag_sp + drag_rw + drag_rwh + drag_fwh + drag_ch

        df_fw = float(df_list[idx["front-wing"]])
        df_sp = float(df_list[idx["sidepod"]])
        df_rw = float(df_list[idx["rear-wing"]])
        df_rwh = float(df_list[idx["rear-wheel"]])
        df_fwh = float(df_list[idx["front-wheel"]])
        df_ch = float(df_list[idx["chassis"]])
        df_net = df_fw + df_sp + df_rw + df_rwh + df_fwh + df_ch

        moment_idx = None  # ex: 8
        moment = float(df_list[moment_idx]) if moment_idx is not None else 0.0

        self.forces = [
        drag_fw,
        drag_sp,
        drag_rw,
        drag_net,
        df_fw,
        df_sp,
        df_rw,
        df_net,
        moment
        ]
        print(self.forces)
        self.write_to_forcesheet()

    def write_to_forcesheet(self):
        source = r"data\aero force sheet.xlsx"
        destination = self.forces_dir
        shutil.copy2(source, destination)
        self.force_book = load_workbook(self.forces_dir / "aero force sheet.xlsx")
        self.force_sheet = self.force_book.active
        '''
            forces must be in following format:
                [Drag Frontwing,
                 Drag Sidepod,
                 Drag Rearwing,
                 Drag Net,
                 Downforce Frontwing,
                 Downforce Sidepod,
                 Downforce Rearwing,
                 Downforce Net,
                 Moment around front axis]
        '''
        assert len(self.forces) == len(rows) == len(cols)

        assert self.force_sheet["A1"].value == "v1.0"

        for i, row in enumerate(rows):
            self.force_sheet.cell(row=rows[i], column=cols[i], value=self.forces[i])
        
        self.force_book.save(self.forces_dir / "aero force sheet.xlsx")
        print(f"Forces saved in: {self.forces_dir}"+r"\aero force sheet.xlsx")
